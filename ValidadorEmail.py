import openpyxl as excel
import threading
import smtplib
import imaplib
import time
import email
import os
import pandas as pd
import numpy as np

mensaje_error = str("No se encontr")

def limpiar_mensajes(user, password, IMAP):
    mail = imaplib.IMAP4_SSL(IMAP)
    mail.login(user, password)
    mail.select("inbox")
    typ, data = mail.search(None, 'ALL')
    for num in data[0].split():
        mail.store(num, '+FLAGS', r'(\Deleted)')
    mail.expunge()
    mail.close()
    mail.logout()
    return

def envia_mensaje_hilo(smtp_host,user,password,bloque_correo,mensaje,):
    server = smtplib.SMTP(smtp_host,587)
    server.starttls()
    server.login(user,password)
    for correo in bloque_correo:
        server.sendmail(user,correo,mensaje)
        print('el correo',correo,'se envio exitosamente')

    server.quit()
    return
    

def enviar_mensaje(correos_por_validar,mensaje,smtp_host,user,password):
    cantidad_correos = len(correos_por_validar)
    bloques_hilos = 5
    ind = 0
    while(ind < cantidad_correos):
        contador = 0
        bloque_correo = []
        while(ind < cantidad_correos and contador < bloques_hilos):
            bloque_correo.append(correos_por_validar[ind])
            ind = ind + 1
            contador = contador + 1
        hilo = threading.Thread(target=envia_mensaje_hilo, args=(smtp_host,user,password,bloque_correo,mensaje,))
        hilo.start()
        time.sleep(0.00001)
    return

def procesa_identificador_correos(data):
    cadena = str(data[0])
    if(cadena == "b''"):
        lista = []
        return lista
    cadena = cadena.replace('b','')
    cadena = cadena.replace("'",'')
    lista = cadena.split(' ')
    return lista

def get_body(msg):
    if(msg.is_multipart()):
        return get_body(msg.get_payload(0))
    else:
        return msg.get_payload(None,True)
    return


def leer_mensajes_no_validos(imap_host,user,password):
    # connect to host using SSL
    imap = imaplib.IMAP4_SSL(imap_host)

    # login to server
    imap.login(user, password)

    imap.select('Inbox')

    tmp, ids = imap.search(None, 'ALL')
    identificador = procesa_identificador_correos(ids)

    correos_no_validos = []
    for idd in identificador:
        nuevo_id = bytes(idd, "utf-8")
        result, data = imap.fetch(nuevo_id,'(RFC822)')
        raw = email.message_from_bytes(data[0][1])
        mensaje_validacion = str(get_body(raw))
        primera_aparicion = mensaje_validacion.find(mensaje_error)
        if(primera_aparicion != -1):
            primera_aparicion += 86
            correo = ""
            while(mensaje_validacion[primera_aparicion] != ' '):
                correo = correo + mensaje_validacion[primera_aparicion]
                primera_aparicion = primera_aparicion + 1
            correos_no_validos.append(correo)
    
    imap.close()
    return correos_no_validos

def recoge_correos(Hoja):
    correos = []
    for row in Hoja.values:
        contador = 0
        for value in row:
            value = str(value)
            if(value == "None"):
                break
            correos.append(value)
            contador = contador + 1
        if(contador == 0):
            break
    return correos

def crea_hoja_resultado(Hoja_Resultado,doc):
    existe = False
    for hojas in doc.sheetnames:
        if(hojas == Hoja_Resultado):
            existe = True
    if(not existe):
        doc.create_sheet(Hoja_Resultado,0)
    doc.save(libro)
    return

def guarda_resultado(columna,correos,Hoja_Resultado,doc):
    if(columna == "A"):
        Hoja_Resultado['A1'] = "Correos Validos"
    else:
        Hoja_Resultado['B1'] = "Correos No Validos"
    for i in range(len(correos)):
        celda = str(columna+str(i+2))
        Hoja_Resultado[celda] = correos[i]
    print(correos)
    doc.save(libro)
    return

if __name__ == '__main__':
    directorio = 'uploads'
    libro = 'prueba' ##debe ser parametro de entrada
    extension = '.xlsx'
    filename = os.path.join(directorio,libro+extension)
    fileoutput = os.path.join(directorio,libro+'_output'+extension)
    #doc = excel.load_workbook(filename)
    #Hoja = doc["BD Correos"]
    df = pd.read_excel(filename,header=None)
    df.columns = ['email']
    print(df.head())
    emails = list(df['email'])

    mensaje = 'Hi, it is message from python'
    user = 'email.python.11@gmail.com'
    password = 'emailpython11'
    smtp_host = 'smtp.gmail.com'
    imap_host = 'imap.gmail.com'

    limpiar_mensajes(user,password,imap_host)

    #correos_por_validar = recoge_correos(emails)

    enviar_mensaje(emails,mensaje,smtp_host,user,password)

    time.sleep(20)
    
    correos_no_validos = leer_mensajes_no_validos(imap_host,user,password)

    resultado = []
    for correo in emails:
        if(correo in correos_no_validos):
            resultado.append('No Valido')
        else:
            resultado.append('Valido')

    print(resultado)
    df['resultado'] = np.array(resultado)
    print(df.head(20))
    df.to_excel(fileoutput)
    #correos_validos = set(emails)

    #for correo in correos_no_validos:
    #    correos_validos.remove(correo)

    #correos_validos = list(correos_validos)
    #print(correos_no_validos)
    #print(correos_validos)
    #crea_hoja_resultado("Resultados",doc)
    #Hoja_Resultado = doc["Resultados"]

    #guarda_resultado('A',correos_validos,Hoja_Resultado,doc)
    #guarda_resultado('B',correos_no_validos,Hoja_Resultado,doc)
    
    
    
    
